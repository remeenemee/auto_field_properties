import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from docx import Document


class AutoFillProperties:
    def __init__(self, master):
        self.master = master
        master.title("AutoFillProperties v0.1")

        self.excel_file_path = ""
        self.word_file_path = ""

        self.button_update = tk.Label(master, text="Автозаполнение полей ПЗ")
        self.button_update.grid(row=0, column=0, columnspan=2, pady=20)

        self.label_excel = tk.Label(
            master, text="Выберите файл базы данных Excel:")
        self.label_excel.grid(row=1, column=0, padx=10, pady=10)

        self.button_excel = tk.Button(
            master, text="Выбрать Excel", command=self.choose_excel_file)
        self.button_excel.grid(row=1, column=1, padx=10, pady=10)

        self.label_word = tk.Label(master, text="Выберите файл Word:")
        self.label_word.grid(row=2, column=0, padx=10, pady=10)

        self.button_word = tk.Button(
            master, text="Выбрать Word", command=self.choose_word_file)
        self.button_word.grid(row=2, column=1, padx=10, pady=10)

        self.button_update = tk.Button(
            master, text="Создать поля документа", command=self.add_standart_custom_properties)
        self.button_update.grid(row=3, column=0, columnspan=2, pady=20)

        self.button_update = tk.Button(
            master, text="Обновить поля из базы данных", command=self.update_properties)
        self.button_update.grid(row=4, column=0, columnspan=2, pady=20)

        self.status_label = tk.Label(master, text="")
        self.status_label.grid(row=5, column=0, columnspan=2)

        self.dev_label = tk.Label(master, text="v0.1 KochevDA")
        self.dev_label.grid(row=6, column=1, columnspan=2)

    def choose_excel_file(self):
        self.excel_file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx;*.xls")])
        self.status_label.config(
            text=f"Выбран файл Excel: {os.path.basename(self.excel_file_path)}")

    def choose_word_file(self):
        self.word_file_path = filedialog.askopenfilename(
            filetypes=[("Word Files", "*.docx")])
        self.status_label.config(
            text=f"Выбран файл Word: {os.path.basename(self.word_file_path)}")

    def add_standart_custom_properties(self):
        word_document = Document(self.word_file_path)

        word_document.custom_properties['наименование_заказчика'] = '-'
        word_document.custom_properties['наименование_объекта'] = '-'
        word_document.custom_properties['год'] = '-'
        word_document.custom_properties['шифр_объекта'] = '-'
        word_document.custom_properties['кадастровый_номер'] = '-'
        word_document.custom_properties['гпзу'] = '-'

        # def add_new_custom_properties():


        word_document.save(self.word_file_path)
        self.status_label.config(text="Свойства документа созданы!")

    def update_properties(self):
        if not self.excel_file_path or not self.word_file_path:
            self.status_label.config(text="Выберите файлы Excel и Word")
            return

        try:
            excel_workbook = load_workbook(self.excel_file_path)
            excel_sheet = excel_workbook.active

            наименование_заказчика = excel_sheet.cell(row=2, column=1).value
            наименование_объекта = excel_sheet.cell(row=2, column=2).value
            год = excel_sheet.cell(row=2, column=3).value
            шифр_объекта = excel_sheet.cell(row=2, column=4).value
            кадастровый_номер = excel_sheet.cell(row=2, column=5).value
            гпзу = excel_sheet.cell(row=2, column=6).value

            word_document = Document(self.word_file_path)

            # Обновление свойств
            if word_document.custom_properties['наименование_заказчика']:
                word_document.custom_properties['наименование_заказчика'] = наименование_заказчика

            if word_document.custom_properties['наименование_объекта']:
                word_document.custom_properties['наименование_объекта'] = наименование_объекта

            if word_document.custom_properties['год']:
                word_document.custom_properties['год'] = год

            if word_document.custom_properties['шифр_объекта']:
                word_document.custom_properties['шифр_объекта'] = шифр_объекта

            if word_document.custom_properties['кадастровый_номер']:
                word_document.custom_properties['кадастровый_номер'] = кадастровый_номер

            if word_document.custom_properties['гпзу']:
                word_document.custom_properties['гпзу'] = гпзу

            word_document.save(self.word_file_path)
            self.status_label.config(
                text="Изменения внесены в свойства документа Word.")

        except Exception as e:
            self.status_label.config(text=f"Ошибка: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = AutoFillProperties(root)
    root.mainloop()
